"""Loads unit spreadsheet for Lease/Rental, owner, customer, and email lookups."""
from pathlib import Path
from datetime import date
import openpyxl

BASE_DIR = Path(__file__).parent
UNIT_FILE = BASE_DIR / "Unit numbers and descriptions  (1).xlsx"
CUSTOMER_FILE = BASE_DIR / "Customer numbers.xlsx"

_unit_data = None
_unit_loaded_date = None
_customer_data = None
_customer_loaded_date = None
_email_data = None
_email_loaded_date = None


def _to_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def _load_units() -> dict:
    global _unit_data, _unit_loaded_date
    today = date.today()
    if _unit_data is not None and _unit_loaded_date == today:
        return _unit_data

    if not UNIT_FILE.exists():
        _unit_data = {}
        return _unit_data

    wb = openpyxl.load_workbook(UNIT_FILE, data_only=True)
    ws = wb.active

    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        unit = row[0]
        if not unit:
            continue
        unit_str = str(unit).strip()
        owner = _to_str(row[2]) if len(row) > 2 else ""
        lease_name = _to_str(row[3]) if len(row) > 3 else ""
        lease_or_rental = lease_name.capitalize() if lease_name.lower() in ("lease", "rental") else "Lease"
        data[unit_str] = {"lease_or_rental": lease_or_rental, "owner": owner}

    _unit_data = data
    _unit_loaded_date = today
    return _unit_data


def _load_emails() -> dict:
    global _email_data, _email_loaded_date
    today = date.today()
    if _email_data is not None and _email_loaded_date == today:
        return _email_data

    _email_data = {}
    _email_loaded_date = today

    if not UNIT_FILE.exists():
        return _email_data

    wb = openpyxl.load_workbook(UNIT_FILE, data_only=True)
    if "Email Addresses" not in wb.sheetnames:
        return _email_data

    ws = wb["Email Addresses"]
    for row in ws.iter_rows(min_row=1, values_only=True):
        key = row[0] if row else None
        email = row[1] if len(row) > 1 else None
        if key and email:
            _email_data[_to_str(key).lower()] = _to_str(email)

    return _email_data


def _load_customers() -> dict:
    global _customer_data, _customer_loaded_date
    today = date.today()
    if _customer_data is not None and _customer_loaded_date == today:
        return _customer_data

    _customer_data = {}
    _customer_loaded_date = today

    if not CUSTOMER_FILE.exists():
        return _customer_data

    wb = openpyxl.load_workbook(CUSTOMER_FILE, data_only=True)
    tabs = [s for s in wb.sheetnames if s.upper() != "TURNED IN"]

    # Iterate most-recent-first; first entry for a unit wins
    for tab_name in reversed(tabs):
        ws = wb[tab_name]
        for row in ws.iter_rows(min_row=4, values_only=True):
            vehicle = row[3] if len(row) > 3 else None
            cust_num = row[1] if len(row) > 1 else None
            cust_name = row[2] if len(row) > 2 else None
            if not vehicle or (not cust_num and not cust_name):
                continue
            key = _to_str(vehicle).lower()
            if key not in _customer_data:
                _customer_data[key] = (_to_str(cust_num), _to_str(cust_name))

    return _customer_data


def lookup_unit(unit_number: str) -> str:
    """Returns 'Lease', 'Rental', or 'Lease' (default) for a given unit number."""
    if not unit_number:
        return "Lease"
    data = _load_units()
    return data.get(str(unit_number).strip(), {}).get("lease_or_rental", "Lease")


def lookup_owner(unit_number: str) -> str:
    """Returns the owner name from column C of the unit numbers spreadsheet."""
    if not unit_number:
        return ""
    data = _load_units()
    return data.get(str(unit_number).strip(), {}).get("owner", "")


def lookup_customer(unit_number: str) -> tuple:
    """Returns (customer_number, customer_name) or ('', '') if not found."""
    if not unit_number:
        return ("", "")
    data = _load_customers()
    return data.get(str(unit_number).strip().lower(), ("", ""))


def lookup_email(unit_number: str, owner: str = "") -> str:
    """Returns email for a unit (or owner name) from the Email Addresses tab."""
    emails = _load_emails()
    if not emails:
        return ""
    unit_key = str(unit_number).strip().lower() if unit_number else ""
    owner_key = str(owner).strip().lower() if owner else ""
    return emails.get(unit_key) or emails.get(owner_key) or ""
