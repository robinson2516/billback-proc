"""Fills the 1 REBILL BLANK.xlsm template with extracted data."""
import io
from pathlib import Path
import openpyxl

BASE_DIR = Path(__file__).parent
TEMPLATE = BASE_DIR / "1 REBILL_BLANK_updated.xlsm"

LINE_ITEM_START_ROW = 15   # line items begin at row 15

_VENDOR_ALIASES = {
    "salt lake city": "Peterbilt of Utah",
    "slc":            "Peterbilt of Utah",
}


def _load_labor_rates(wb) -> dict:
    """Return {vendor_name_lower: (current_rate, pac_rate)} from LaborRates sheet."""
    rates = {}
    if "LaborRates" not in wb.sheetnames:
        return rates
    lr = wb["LaborRates"]
    for row in lr.iter_rows(min_row=2, values_only=True):
        vendor, _, current, pac = (row[i] if i < len(row) else None for i in range(4))
        if vendor and current and pac:
            try:
                rates[str(vendor).strip().lower()] = (float(current), float(pac))
            except (ValueError, TypeError):
                continue
    return rates


def _match_vendor(name: str, labor_rates: dict) -> str:
    """Return the exact LaborRates vendor string closest to `name`, or `name` if no match."""
    if not name:
        return name
    normalized = name.strip().lower()
    # Check aliases first
    if normalized in _VENDOR_ALIASES:
        normalized = _VENDOR_ALIASES[normalized].lower()
    # Exact match against LaborRates keys
    for vendor_key in labor_rates:
        if vendor_key == normalized:
            return vendor_key.title().replace("Petebilt", "Petebilt")
    # City-name match: input may be just the city (e.g. "Elko" → "Elko Peterbilt")
    for vendor_key in labor_rates:
        words = vendor_key.split()
        if normalized in words:
            return vendor_key  # return lowercase key; used only for rate lookup
    # Partial word overlap (longest match wins)
    input_words = set(normalized.split()) - {"peterbilt", "petebilt", "of", "the"}
    best, best_score = None, 0
    for vendor_key in labor_rates:
        vendor_words = set(vendor_key.split()) - {"peterbilt", "petebilt", "of", "the"}
        score = len(input_words & vendor_words)
        if score > best_score:
            best, best_score = vendor_key, score
    return best if best else name


def fill_rebill_sheet(data: dict) -> bytes:
    wb = openpyxl.load_workbook(TEMPLATE, keep_vba=True)
    ws = wb["Rebill Sheet"]

    # Force Excel to recalculate all formulas on open
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    # Set sheet zoom to 100%
    ws.sheet_view.zoomScale = 100

    # ── Vendor & labor rates ──────────────────────────────────────
    labor_rates = _load_labor_rates(wb)
    vendor_key  = _match_vendor(data.get("vendor") or "", labor_rates)
    ws.cell(row=1, column=4).value = vendor_key
    current_rate, pac_rate = labor_rates.get(vendor_key.strip().lower(), (0.0, 0.0))

    # ── Header fields ─────────────────────────────────────────────
    ws.cell(row=1, column=17).value = data.get("customer") or ""
    ws.cell(row=2, column=4).value  = data.get("invoice_number") or ""
    ws.cell(row=2, column=17).value = data.get("taxable") or "Yes"
    ws.cell(row=3, column=4).value  = data.get("date") or ""
    ws.cell(row=4, column=4).value  = data.get("lease_or_rental") or "Lease"
    ws.cell(row=4, column=15).value = data.get("invoice_wording") or ""

    # ── Misc / shop supplies → K9 ─────────────────────────────────
    try:
        misc = float(data.get("misc") or 0)
    except (ValueError, TypeError):
        misc = 0.0
    ws.cell(row=9, column=11).value = misc if misc else None

    # ── Line items ────────────────────────────────────────────────
    # Column mapping:
    #   Col B (2)  — Rebill Parts
    #   Col C (3)  — Rebill Labor
    #   Col D (4)  — Internal PMs
    #   Col E (5)  — Internal Repairs
    # For rebill labor rows, also populate Q/R/S (hours, rate, total)
    line_items = data.get("line_items", [])
    total = 0.0
    row = LINE_ITEM_START_ROW
    for item in line_items:
        try:
            cost = float(item.get("cost") or 0)
        except (ValueError, TypeError):
            cost = 0.0
        if cost == 0:
            continue

        total += cost
        item_type     = (item.get("type")     or "repair").lower()
        item_category = (item.get("category") or "parts").lower()

        if item_type == "pm":
            ws.cell(row=row, column=4).value = cost
        elif item_type == "rebill":
            if item_category == "labor":
                ws.cell(row=row, column=3).value = cost
                # Calculate and write Q (hours), R (rate), S (total)
                if pac_rate and current_rate:
                    hours     = round(cost / pac_rate, 2)
                    adj_rate  = round(current_rate - 10, 2)
                    lab_total = round(hours * adj_rate, 2)
                    ws.cell(row=row, column=17).value = hours
                    ws.cell(row=row, column=18).value = adj_rate
                    ws.cell(row=row, column=19).value = lab_total
            else:
                ws.cell(row=row, column=2).value = cost
        else:  # repair
            ws.cell(row=row, column=5).value = cost

        row += 1

    # ── Invoice total (line items + misc) → L12 ──────────────────
    inv_total = total + misc
    ws.cell(row=12, column=12).value = round(inv_total, 2) if inv_total else None

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
